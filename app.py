# -*- coding: utf-8 -*-
import base64
import hashlib
import json
import math
import os
import re
import sqlite3
import threading
import time
import threading as _threading
from datetime import date, datetime, timezone, timedelta

from flask import Flask, g, jsonify, redirect, render_template, request, Response, session
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('DB_PATH', os.path.join(BASE_DIR, 'gps.db'))
DATABASE_URL = os.environ.get('DATABASE_URL', '')
PG = DATABASE_URL.startswith('postgres')
VENDORS_FILE = os.path.join(BASE_DIR, 'vendors.json')
PDV_FILE = os.path.join(BASE_DIR, 'pdv.json')
RUTAS_FILE = os.path.join(BASE_DIR, 'rutas.json')
OVERRIDES_FILE = os.path.join(BASE_DIR, 'overrides.json')

APP_VERSION = '3.1'
APP_VERSION_CODE = 12
APK_URL = 'https://github.com/GonzaloBorras/gps-vendedores/releases/download/apk-v1.0/GPS-Merchan.apk'
APK_URL_ADMIN = 'https://github.com/GonzaloBorras/gps-vendedores/releases/download/apk-admin/GPS-Admin.apk'

with open(VENDORS_FILE, encoding='utf-8') as f:
    VENDORS = json.load(f)

OVERRIDES = {}
if os.path.exists(OVERRIDES_FILE):
    try:
        with open(OVERRIDES_FILE, encoding='utf-8') as f:
            OVERRIDES = json.load(f)
    except Exception:
        logging.getLogger(__name__).error('Failed to load overrides file')
        OVERRIDES = {}

for v in VENDORS:
    o = OVERRIDES.get(v['code'])
    if o:
        v['name'] = o.get('name', v['name'])
        v['color'] = o.get('color', v['color'])

PDV = []
if os.path.exists(PDV_FILE):
    with open(PDV_FILE, encoding='utf-8') as f:
        PDV = json.load(f)

PDV_BY_CODE = {p['c']: p for p in PDV}

_PDV_LOCK = _threading.Lock()

def _pdv_snapshot():
    with _PDV_LOCK:
        return list(PDV)

RUTAS = {}
if os.path.exists(RUTAS_FILE):
    with open(RUTAS_FILE, encoding='utf-8') as f:
        RUTAS = json.load(f)

VENDOR_BY_CODE = {v['code'].upper(): v for v in VENDORS}

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cambiar-esta-clave-por-una-segura')
DASH_PIN = os.environ.get('DASH_PIN', '1234')

class _RateLimiter:
    def __init__(self):
        self._hits = {}
        self._lock = _threading.Lock()
    
    def check(self, key, limit, window_s=60):
        now = time.time()
        with self._lock:
            hits = self._hits.get(key, [])
            hits = [t for t in hits if now - t < window_s]
            if len(hits) >= limit:
                return False
            hits.append(now)
            self._hits[key] = hits
            return True

_rate_limiter = _RateLimiter()

ACTIVE_MS = 120000  # 2 minutos: se considera activo si mandó posición hace menos que esto
VISIT_RADIUS_M = 150  # radio por defecto para validar que el merchan está en el PDV
START_RADIUS_M = 400  # radio para permitir iniciar en vivo (tolerancia a coords aproximadas)
VISIT_MAX_MS = 5 * 60 * 1000  # la posición para validar la visita no puede tener más de 5 min
ACTIVE_GAP_MS = 5 * 60 * 1000  # para reportes: salto de más de 5 min no cuenta como actividad
SHIFT_DEFAULT_MS = 8 * 60 * 60 * 1000  # jornada por defecto: 8 hs
VAPID_MAILTO = 'mailto:admin@gps-merchan.com'

RUTAS_DIA_NAMES = {
    'LUNES': 'Lunes', 'MARTES': 'Martes', 'MIERCOLES': 'Miércoles',
    'JUEVES': 'Jueves', 'VIERNES': 'Viernes', 'SABADO': 'Sábado', 'DOMINGO': 'Domingo',
}
WEEKDAY_KEYS = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']


# ---------------- Conexión a base de datos (SQLite local / PostgreSQL en Render) ----------------

def _q(sql):
    return sql.replace('?', '%s') if PG else sql


def _raw_conn():
    if PG:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def get_db():
    db = getattr(g, '_db', None)
    if db is not None:
        return db
    if PG:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        db = g._db = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    else:
        db = g._db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        db.execute('PRAGMA journal_mode=WAL')
        db.execute('PRAGMA busy_timeout=5000')
    return db


@app.teardown_appcontext
def close_db(exc=None):
    db = getattr(g, '_db', None)
    if db is not None:
        db.close()


def _exec(con, sql, params=None):
    if PG:
        cur = con.cursor()
        cur.execute(_q(sql), params or ())
        return cur
    return con.execute(_q(sql), params or ())


def _haversine(lat1, lon1, lat2, lon2):
    R = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


def _today_str():
    return (datetime.now(timezone.utc) - timedelta(hours=3)).strftime('%Y-%m-%d')


def _today_key():
    wd = (datetime.now(timezone.utc) - timedelta(hours=3)).weekday()
    return WEEKDAY_KEYS[wd]


def _day_range(fecha):
    start = datetime.strptime(fecha, '%Y-%m-%d').replace(tzinfo=timezone.utc) + timedelta(hours=3)
    start_ms = int(start.timestamp() * 1000)
    return start_ms, start_ms + 86400000


JORNADA_START_HM = 8 * 60       # 08:00
JORNADA_END_HM_LV = 16 * 60     # 16:00 de lunes a viernes
JORNADA_END_HM_SAB = 13 * 60    # 13:00 los sábados


def _jornada_window(code):
    """Ventana (start_ms, end_ms) de hoy para el merchan en hora Argentina.
    None en domingo. El final se acota con las horas configuradas por merchan."""
    now = datetime.now(timezone(timedelta(hours=-3)))
    wd = now.weekday()  # 0=Lu ... 5=Sá, 6=Do
    if wd == 6:
        return None
    end_hm = JORNADA_END_HM_SAB if wd == 5 else JORNADA_END_HM_LV
    day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_ts = int((day + timedelta(minutes=JORNADA_START_HM)).timestamp() * 1000)
    end_ts = int((day + timedelta(minutes=end_hm)).timestamp() * 1000)
    srow = _exec(get_db(), 'SELECT shift_ms FROM shifts WHERE code=?', (code,)).fetchone()
    if srow:
        end_ts = min(end_ts, start_ts + srow['shift_ms'])
    return start_ts, end_ts


# ---------------- Esquema ----------------

def init_db():
    con = _raw_conn()
    if PG:
        cur = con.cursor()
        for s in [
            'CREATE TABLE IF NOT EXISTS vendors (code TEXT PRIMARY KEY, name TEXT NOT NULL, prov TEXT NOT NULL, color TEXT NOT NULL, grupo TEXT NOT NULL DEFAULT \'rutas\')',
            'CREATE TABLE IF NOT EXISTS positions (id SERIAL PRIMARY KEY, code TEXT NOT NULL, name TEXT NOT NULL, lat DOUBLE PRECISION NOT NULL, lon DOUBLE PRECISION NOT NULL, session TEXT, ts BIGINT NOT NULL)',
            'CREATE INDEX IF NOT EXISTS idx_positions_code_ts ON positions(code, ts)',
            'CREATE TABLE IF NOT EXISTS visitas (fecha TEXT NOT NULL, code TEXT NOT NULL, cliente TEXT NOT NULL, razon TEXT NOT NULL DEFAULT \'\', calle TEXT NOT NULL DEFAULT \'\', vta TEXT NOT NULL DEFAULT \'\', ts BIGINT NOT NULL, foto TEXT, foto_ts BIGINT, foto_salida_ts BIGINT, PRIMARY KEY (fecha, code, cliente))',
            'CREATE INDEX IF NOT EXISTS idx_visitas_fecha_code ON visitas(fecha, code)',
            'CREATE TABLE IF NOT EXISTS alerts (code TEXT PRIMARY KEY, tipo TEXT NOT NULL DEFAULT \'gps_off\', ts BIGINT NOT NULL, msj TEXT NOT NULL DEFAULT \'\')',
            'CREATE TABLE IF NOT EXISTS shifts (code TEXT PRIMARY KEY, shift_ms BIGINT NOT NULL)',
            'CREATE TABLE IF NOT EXISTS absence (fecha TEXT NOT NULL, code TEXT NOT NULL, motivo TEXT NOT NULL DEFAULT \'\', ts BIGINT NOT NULL, PRIMARY KEY (fecha, code))',
            'CREATE TABLE IF NOT EXISTS messages (id SERIAL PRIMARY KEY, code TEXT NOT NULL, msj TEXT NOT NULL, ts BIGINT NOT NULL, visto INTEGER NOT NULL DEFAULT 0)',
            'CREATE INDEX IF NOT EXISTS idx_messages_code ON messages(code)',
            'CREATE TABLE IF NOT EXISTS pdv_radius (cliente TEXT PRIMARY KEY, radius_m INTEGER NOT NULL)',
            'CREATE TABLE IF NOT EXISTS pdvs_extra (cliente TEXT PRIMARY KEY, razon TEXT NOT NULL, calle TEXT NOT NULL DEFAULT \'\', altura TEXT NOT NULL DEFAULT \'\', vta TEXT NOT NULL DEFAULT \'\', prov TEXT NOT NULL DEFAULT \'\', telefono TEXT NOT NULL DEFAULT \'\', contacto TEXT NOT NULL DEFAULT \'\', notas TEXT NOT NULL DEFAULT \'\', lat DOUBLE PRECISION NOT NULL, lon DOUBLE PRECISION NOT NULL, creado_por TEXT NOT NULL DEFAULT \'\', ts BIGINT NOT NULL)',
            'CREATE TABLE IF NOT EXISTS geoevents (id SERIAL PRIMARY KEY, code TEXT NOT NULL, cliente TEXT NOT NULL DEFAULT \'\', razon TEXT NOT NULL DEFAULT \'\', tipo TEXT NOT NULL DEFAULT \'\', dist_m INTEGER, ts BIGINT NOT NULL)',
            'CREATE INDEX IF NOT EXISTS idx_geoevents_code_ts ON geoevents(code, ts)',
            'CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)',
            'CREATE TABLE IF NOT EXISTS push_subs (endpoint TEXT PRIMARY KEY, code TEXT NOT NULL, p256dh TEXT NOT NULL, auth TEXT NOT NULL, ts BIGINT NOT NULL)',
            'CREATE INDEX IF NOT EXISTS idx_push_subs_code ON push_subs(code)',
            'CREATE TABLE IF NOT EXISTS devices (code TEXT PRIMARY KEY, app TEXT NOT NULL DEFAULT \'web\', app_version TEXT NOT NULL DEFAULT \'\', version_code INTEGER NOT NULL DEFAULT 0, updated_at BIGINT NOT NULL)',
        ]:
            cur.execute(s)
        cur.execute('ALTER TABLE visitas ADD COLUMN IF NOT EXISTS foto TEXT')
        cur.execute('ALTER TABLE visitas ADD COLUMN IF NOT EXISTS foto_ts BIGINT')
        cur.execute('ALTER TABLE visitas ADD COLUMN IF NOT EXISTS foto_salida_ts BIGINT')
        cur.execute('ALTER TABLE positions ADD COLUMN IF NOT EXISTS battery INTEGER')
        cur.execute('ALTER TABLE pdvs_extra ADD COLUMN IF NOT EXISTS telefono TEXT NOT NULL DEFAULT \'\'')
        cur.execute('ALTER TABLE pdvs_extra ADD COLUMN IF NOT EXISTS contacto TEXT NOT NULL DEFAULT \'\'')
        cur.execute('ALTER TABLE pdvs_extra ADD COLUMN IF NOT EXISTS notas TEXT NOT NULL DEFAULT \'\'')
        con.commit()
    else:
        con.executescript('''
            CREATE TABLE IF NOT EXISTS vendors (
                code  TEXT PRIMARY KEY,
                name  TEXT NOT NULL,
                prov  TEXT NOT NULL,
                color TEXT NOT NULL,
                grupo TEXT NOT NULL DEFAULT 'rutas'
            );
            CREATE TABLE IF NOT EXISTS positions (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                code    TEXT NOT NULL,
                name    TEXT NOT NULL,
                lat     REAL NOT NULL,
                lon     REAL NOT NULL,
                session TEXT,
                ts      INTEGER NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_positions_code_ts ON positions(code, ts);
            CREATE TABLE IF NOT EXISTS visitas (
                fecha   TEXT NOT NULL,
                code    TEXT NOT NULL,
                cliente TEXT NOT NULL,
                razon   TEXT NOT NULL DEFAULT '',
                calle   TEXT NOT NULL DEFAULT '',
                vta     TEXT NOT NULL DEFAULT '',
                ts      INTEGER NOT NULL,
                foto    TEXT,
                foto_ts INTEGER,
                foto_salida_ts INTEGER,
                PRIMARY KEY (fecha, code, cliente)
            );
            CREATE INDEX IF NOT EXISTS idx_visitas_fecha_code ON visitas(fecha, code);
            CREATE TABLE IF NOT EXISTS alerts (
                code TEXT PRIMARY KEY,
                tipo TEXT NOT NULL DEFAULT 'gps_off',
                ts INTEGER NOT NULL,
                msj TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS shifts (
                code TEXT PRIMARY KEY,
                shift_ms INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS absence (
                fecha TEXT NOT NULL,
                code TEXT NOT NULL,
                motivo TEXT NOT NULL DEFAULT '',
                ts INTEGER NOT NULL,
                PRIMARY KEY (fecha, code)
            );
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                msj TEXT NOT NULL,
                ts INTEGER NOT NULL,
                visto INTEGER NOT NULL DEFAULT 0
            );
            CREATE INDEX IF NOT EXISTS idx_messages_code ON messages(code);
            CREATE TABLE IF NOT EXISTS pdv_radius (
                cliente TEXT PRIMARY KEY,
                radius_m INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS pdvs_extra (
                cliente    TEXT PRIMARY KEY,
                razon      TEXT NOT NULL,
                calle      TEXT NOT NULL DEFAULT '',
                altura     TEXT NOT NULL DEFAULT '',
                vta        TEXT NOT NULL DEFAULT '',
                prov       TEXT NOT NULL DEFAULT '',
                telefono   TEXT NOT NULL DEFAULT '',
                contacto   TEXT NOT NULL DEFAULT '',
                notas      TEXT NOT NULL DEFAULT '',
                lat        REAL NOT NULL,
                lon        REAL NOT NULL,
                creado_por TEXT NOT NULL DEFAULT '',
                ts         INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS geoevents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                cliente TEXT NOT NULL DEFAULT '',
                razon TEXT NOT NULL DEFAULT '',
                tipo TEXT NOT NULL DEFAULT '',
                dist_m INTEGER,
                ts INTEGER NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_geoevents_code_ts ON geoevents(code, ts);
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );
            CREATE TABLE IF NOT EXISTS push_subs (
                endpoint TEXT PRIMARY KEY,
                code TEXT NOT NULL,
                p256dh TEXT NOT NULL,
                auth TEXT NOT NULL,
                ts INTEGER NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_push_subs_code ON push_subs(code);
            CREATE TABLE IF NOT EXISTS devices (
                code TEXT PRIMARY KEY,
                app TEXT NOT NULL DEFAULT 'web',
                app_version TEXT NOT NULL DEFAULT '',
                version_code INTEGER NOT NULL DEFAULT 0,
                updated_at INTEGER NOT NULL
            );
        ''')
        vcols = [r[1] for r in con.execute('PRAGMA table_info(visitas)')]
        if 'foto' not in vcols:
            con.execute('ALTER TABLE visitas ADD COLUMN foto TEXT')
        if 'foto_ts' not in vcols:
            con.execute('ALTER TABLE visitas ADD COLUMN foto_ts INTEGER')
        if 'foto_salida_ts' not in vcols:
            con.execute('ALTER TABLE visitas ADD COLUMN foto_salida_ts INTEGER')
        vcols2 = [r[1] for r in con.execute('PRAGMA table_info(vendors)')]
        if 'grupo' not in vcols2:
            con.execute("ALTER TABLE vendors ADD COLUMN grupo TEXT NOT NULL DEFAULT 'rutas'")
        pcols = [r[1] for r in con.execute('PRAGMA table_info(pdvs_extra)')]
        if 'telefono' not in pcols:
            con.execute("ALTER TABLE pdvs_extra ADD COLUMN telefono TEXT NOT NULL DEFAULT ''")
        if 'contacto' not in pcols:
            con.execute("ALTER TABLE pdvs_extra ADD COLUMN contacto TEXT NOT NULL DEFAULT ''")
        if 'notas' not in pcols:
            con.execute("ALTER TABLE pdvs_extra ADD COLUMN notas TEXT NOT NULL DEFAULT ''")
        pcolsb = [r[1] for r in con.execute('PRAGMA table_info(positions)')]
        if 'battery' not in pcolsb:
            con.execute('ALTER TABLE positions ADD COLUMN battery INTEGER')

    _exec(con, 'DELETE FROM vendors')
    for v in VENDORS:
        _exec(con, 'INSERT INTO vendors(code, name, prov, color, grupo) VALUES (?,?,?,?,?)',
              (v['code'], v['name'], v['prov'], v['color'], v.get('grupo', 'rutas')))
    con.commit()
    con.close()
    _refresh_pdv_radius()


_PDV_RADIUS = {}


def _refresh_pdv_radius():
    global _PDV_RADIUS
    try:
        con = _raw_conn()
        _PDV_RADIUS = {r['cliente']: r['radius_m'] for r in _exec(con, 'SELECT cliente, radius_m FROM pdv_radius')}
        con.close()
    except Exception:
        logging.getLogger(__name__).error('Failed to refresh PDV radius')
        _PDV_RADIUS = {}


def _radius_for(cliente):
    return _PDV_RADIUS.get(str(cliente), VISIT_RADIUS_M)


init_db()


# ---------------- Limpieza automática de datos (fotos y recorrido) ----------------

CLEANUP_DAYS = 15  # retención del recorrido en días (posiciones/geoevents)
CLEANUP_INTERVAL_S = 3600  # revisar cada hora


def _ar_now():
    return datetime.now(timezone.utc) - timedelta(hours=3)


def _cleanup_rolling():
    con = _raw_conn()
    try:
        now_ar = _ar_now()
        cutoff_ms = int((now_ar - timedelta(days=CLEANUP_DAYS)).timestamp() * 1000)
        today_ar = now_ar.date().isoformat()
        _exec(con, _q('DELETE FROM positions WHERE ts < ?'), (cutoff_ms,))
        _exec(con, _q('DELETE FROM geoevents WHERE ts < ?'), (cutoff_ms,))
        _exec(con, _q('UPDATE visitas SET foto = NULL, foto_ts = NULL WHERE fecha < ?'), (today_ar,))
        con.commit()
    finally:
        con.close()


def _cleanup_monthly():
    now_ar = _ar_now()
    if now_ar.day > 3:
        return
    month_key = now_ar.strftime('%Y-%m')
    con = _raw_conn()
    try:
        row = _exec(con, "SELECT value FROM settings WHERE key='last_monthly_cleanup'").fetchone()
        if row and row['value'] == month_key:
            return
        first = now_ar.date().replace(day=1)
        start_ms, _ = _day_range(first.isoformat())
        _exec(con, _q('DELETE FROM positions WHERE ts < ?'), (start_ms,))
        _exec(con, _q('DELETE FROM geoevents WHERE ts < ?'), (start_ms,))
        _exec(con, _q('UPDATE visitas SET foto = NULL, foto_ts = NULL WHERE fecha < ?'), (first.isoformat(),))
        _exec(con, '''INSERT INTO settings(key, value) VALUES ('last_monthly_cleanup', ?)
                     ON CONFLICT(key) DO UPDATE SET value = excluded.value''', (month_key,))
        con.commit()
    finally:
        con.close()


def _cleanup_loop():
    while True:
        try:
            _cleanup_rolling()
            _cleanup_monthly()
        except Exception:
            logging.getLogger(__name__).error('Cleanup loop error')
        time.sleep(CLEANUP_INTERVAL_S)


_cleanup_rolling()
threading.Thread(target=_cleanup_loop, daemon=True).start()


def _db_sizes():
    out = {'ok': True, 'tablas': []}
    con = _raw_conn()
    try:
        if PG:
            rows = _exec(con, """SELECT c.relname AS tabla, pg_total_relation_size(c.oid) AS bytes,
                              COALESCE(s.n_live_tup, 0) AS filas
                              FROM pg_class c
                              JOIN pg_namespace n ON n.oid = c.relnamespace
                              LEFT JOIN pg_stat_user_tables s ON s.relid = c.oid
                              WHERE n.nspname = 'public' AND c.relkind = 'r'
                              ORDER BY bytes DESC""").fetchall()
        else:
            rows = _exec(con, """SELECT name AS tabla, (SELECT SUM(pgsize) FROM dbstat WHERE name = m.name)
                              AS bytes, 0 AS filas
                              FROM sqlite_master m WHERE type = 'table' ORDER BY bytes DESC""").fetchall()
        for r in rows:
            out['tablas'].append(dict(r))
    finally:
        con.close()
    return out


MAINT_STATE = {}


@app.route('/api/maintenance')
def maintenance():
    if not (session.get('auth') or request.args.get('pin') == DASH_PIN):
        return jsonify({'ok': False, 'error': 'no autorizado'}), 401
    info = {}
    try:
        if request.args.get('recompress') and not request.args.get('status'):
            if MAINT_STATE.get('running'):
                return jsonify({'ok': True, 'recompress_running': True, **MAINT_STATE})
            MAINT_STATE.clear()
            MAINT_STATE.update({'running': True, 'done': 0, 'total': 0, 'bytes_antes': 0,
                                'bytes_despues': 0, 'por_fecha': {}})
            threading.Thread(target=_recompress_bg, daemon=True).start()
            return jsonify({'ok': True, 'recompress_running': True, 'started': True})
        if request.args.get('status'):
            return jsonify({'ok': True, 'recompress_running': MAINT_STATE.get('running', False),
                            **MAINT_STATE})
        delc = request.args.get('del_pdv', '').strip()
        if delc:
            db = get_db()
            _exec(db, 'DELETE FROM pdvs_extra WHERE cliente = ?', (delc,))
            db.commit()
            PDV_BY_CODE.pop(delc, None)
            with _PDV_LOCK:
                for lst in (PDV, PDV_EXTRA):
                    for p in list(lst):
                        if p.get('c') == delc:
                            lst.remove(p)
            info['del_pdv'] = delc
        if request.args.get('purge'):
            _cleanup_rolling()
            _cleanup_monthly()
        if request.args.get('vacuum') and PG:
            import psycopg2
            con = psycopg2.connect(DATABASE_URL)
            con.autocommit = True
            try:
                cur = con.cursor()
                cur.execute('VACUUM (FULL) positions')
                cur.execute('VACUUM (FULL) visitas')
                cur.execute('VACUUM (FULL) geoevents')
                cur.execute('VACUUM (FULL) messages')
            finally:
                con.close()
        out = _db_sizes()
        out.update(info)
        if MAINT_STATE:
            out['maint'] = {k: v for k, v in MAINT_STATE.items()}
        out['backup_url'] = '/api/maintenance/backup?pin=' + DASH_PIN
        return jsonify(out)
    except Exception as e:
        app.logger.error('Maintenance error: %s', e)
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/maintenance/backup')
def backup_db():
    if not session.get('auth'):
        pin = request.args.get('pin', '').strip()
        if pin != DASH_PIN:
            return jsonify({'ok': False, 'error': 'auth required'}), 401
    import io, gzip
    db = get_db()
    tables = ['visitas', 'pdvs_extra', 'pdv_radius', 'geoevents', 'shifts', 'absence', 'messages', 'push_subs', 'devices', 'settings']
    backup = {}
    for t in tables:
        try:
            rows = _exec(db, 'SELECT * FROM ' + t).fetchall()
            backup[t] = [dict(r) for r in rows]
        except Exception as e:
            backup[t] = {'error': str(e)}
    backup['_meta'] = {'fecha': _today_str(), 'ts': int(time.time() * 1000), 'tables': len(backup)}
    raw = json.dumps(backup, ensure_ascii=False, default=str).encode('utf-8')
    buf = io.BytesIO()
    with gzip.GzipFile(fileobj=buf, mode='wb') as gz:
        gz.write(raw)
    compressed = buf.getvalue()
    return Response(compressed, mimetype='application/gzip',
                    headers={
                        'Content-Disposition': 'attachment; filename=backup_%s.json.gz' % _today_str(),
                        'Content-Length': str(len(compressed))
                    })


def _recompress_bg():
    import io
    from PIL import Image
    con = _raw_conn()
    done = 0
    skipped = 0
    total_after = 0
    try:
        rows = _exec(con, 'SELECT fecha, code, cliente, foto FROM visitas WHERE foto IS NOT NULL').fetchall()
        MAINT_STATE['total'] = len(rows)
        for r in rows:
            try:
                raw = base64.b64decode(r['foto'])
            except Exception:
                logging.getLogger(__name__).error('Failed to decode photo')
                skipped += 1
                continue
            MAINT_STATE['bytes_antes'] += len(raw)
            f = r['fecha']
            MAINT_STATE['por_fecha'][str(f)] = MAINT_STATE['por_fecha'].get(str(f), 0) + 1
            if len(raw) < 300000:
                total_after += len(raw)
                skipped += 1
                MAINT_STATE['done'] = done + skipped
                continue
            try:
                im = Image.open(io.BytesIO(raw))
                im.load()
                im = im.convert('RGB')
                im.thumbnail((1024, 1024))
                buf = io.BytesIO()
                im.save(buf, format='JPEG', quality=72)
                new_b64 = base64.b64encode(buf.getvalue()).decode('ascii')
                _exec(con, _q('UPDATE visitas SET foto = ? WHERE fecha = ? AND code = ? AND cliente = ?'),
                      (new_b64, r['fecha'], r['code'], r['cliente']))
                total_after += len(new_b64.encode('ascii'))
                done += 1
            except Exception:
                logging.getLogger(__name__).error('Failed to recompress photo')
                total_after += len(raw)
                skipped += 1
            MAINT_STATE['done'] = done + skipped
        con.commit()
    except Exception as e:
        logging.getLogger(__name__).error('Recompress error: %s', e)
        MAINT_STATE['error'] = str(e)
    finally:
        try:
            con.close()
        except Exception:
            logging.getLogger(__name__).error('Failed to close DB connection')
        MAINT_STATE['running'] = False
        MAINT_STATE['fotos'] = done
        MAINT_STATE['sin_cambios'] = skipped
        MAINT_STATE['bytes_despues'] = total_after


# ---------------- PDV adicionales (creados por los merchans desde el mapa) ----------------

def _load_pdv_extra():
    extra = []
    try:
        con = _raw_conn()
        rows = _exec(con, 'SELECT cliente, razon, calle, altura, vta, prov, telefono, contacto, notas, lat, lon FROM pdvs_extra').fetchall()
        con.close()
        for r in rows:
            extra.append({'c': r['cliente'], 'r': r['razon'], 'calle': r['calle'], 'altura': r['altura'],
                          'vta': r['vta'], 'prov': r['prov'], 'telefono': r['telefono'],
                          'contacto': r['contacto'], 'notas': r['notas'],
                          'lat': r['lat'], 'lon': r['lon']})
    except Exception:
        logging.getLogger(__name__).error('Failed to load PDV extras')
    return extra


PDV_EXTRA = _load_pdv_extra()
with _PDV_LOCK:
    PDV.extend(PDV_EXTRA)
    PDV_BY_CODE.update({p['c']: p for p in PDV_EXTRA})


# ---------------- Geocercas ----------------

def check_geofence(db, code, lat, lon, now):
    items = RUTAS.get(code, {}).get(_today_key())
    if not items:
        return
    near = []
    for cliente, _fb in items:
        p = PDV_BY_CODE.get(cliente)
        if not p:
            continue
        pl, po = p.get('lat'), p.get('lon')
        if not pl or not po or (pl == 0 and po == 0):
            continue
        d = _haversine(lat, lon, pl, po)
        if d <= _radius_for(cliente):
            near.append((d, cliente, p.get('r', '')))
    near.sort(key=lambda x: x[0])
    last = _exec(db, 'SELECT tipo, cliente, razon, ts FROM geoevents WHERE code=? ORDER BY ts DESC LIMIT 1',
                 (code,)).fetchone()
    if near:
        d, cliente, razon = near[0]
        if last is None or last['tipo'] != 'enter' or last['cliente'] != cliente or now - last['ts'] > 30000:
            _exec(db, 'INSERT INTO geoevents(code, cliente, razon, tipo, dist_m, ts) VALUES (?,?,?,?,?,?)',
                  (code, cliente, razon, 'enter', int(d), now))
    elif last is not None and last['tipo'] == 'enter' and now - last['ts'] > 10000:
        _exec(db, 'INSERT INTO geoevents(code, cliente, razon, tipo, dist_m, ts) VALUES (?,?,?,?,?,?)',
              (code, last['cliente'], last['razon'], 'leave', None, now))


# ---------------- API ----------------

def _upsert_device(db, code, app, version, version_code):
    try:
        vc = int(version_code)
    except (TypeError, ValueError):
        vc = 0
    v = str(version)[:20] if version else ''
    a = str(app)[:20] if app else 'web'
    if a == 'web':
        row = _exec(db, 'SELECT app FROM devices WHERE code=?', (code,)).fetchone()
        if row and row['app'] == 'android':
            return
    _exec(db, '''INSERT INTO devices(code, app, app_version, version_code, updated_at) VALUES (?,?,?,?,?)
                 ON CONFLICT(code) DO UPDATE SET app = excluded.app, app_version = excluded.app_version,
                 version_code = excluded.version_code, updated_at = excluded.updated_at''',
          (code, a, v, vc, int(time.time() * 1000)))


@app.route('/api/app-version')
def app_version():
    kind = request.args.get('app', 'merchan')
    return jsonify({
        'version': APP_VERSION,
        'versionCode': APP_VERSION_CODE,
        'latestVersion': APP_VERSION,
        'latestVersionCode': APP_VERSION_CODE,
        'apkUrl': APK_URL_ADMIN if kind == 'admin' else APK_URL,
        'notes': ''
    })


@app.route('/api/track', methods=['POST'])
def track():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    rate_code = code or 'anon'
    if not _rate_limiter.check('track:' + rate_code, 30):
        return jsonify({'ok': False, 'error': 'rate limit'}), 429
    vendor = VENDOR_BY_CODE.get(code)
    if not vendor:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 403
    app.logger.info('track: %s at %s,%s', code, body.get('lat'), body.get('lon'))
    try:
        lat = float(body.get('lat'))
        lon = float(body.get('lon'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'coordenadas invalidas'}), 400
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return jsonify({'ok': False, 'error': 'coordenadas fuera de rango'}), 400

    db = get_db()
    ts = int(time.time() * 1000)
    _upsert_device(db, code, body.get('app'), body.get('version'), body.get('versionCode'))
    # Corte de jornada por horario fijo (08:00-16:00 de lunes a viernes,
    # 08:00-13:00 los sábados, domingo sin jornada) para todos los merchans.
    win = _jornada_window(code)
    if not win or ts < win[0] or ts >= win[1]:
        _exec(db, 'DELETE FROM alerts WHERE code = ?', (code,))
        db.commit()
        return jsonify({'ok': True, 'ts': ts, 'jornada_fin': True}), 200
    batt = body.get('battery')
    try:
        batt = int(batt) if batt is not None else None
    except (TypeError, ValueError):
        batt = None
    _exec(db, 'INSERT INTO positions(code, name, lat, lon, session, ts, battery) VALUES (?,?,?,?,?,?,?)',
          (code, vendor['name'], lat, lon, str(body.get('session'))[:64], ts, batt))
    _exec(db, 'DELETE FROM alerts WHERE code = ?', (code,))
    check_geofence(db, code, lat, lon, ts)
    db.commit()
    return jsonify({'ok': True, 'ts': ts})


@app.route('/api/gps-status', methods=['POST'])
def gps_status():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    db = get_db()
    ts = int(time.time() * 1000)
    name = VENDOR_BY_CODE[code]['name']
    _upsert_device(db, code, body.get('app'), body.get('version'), body.get('versionCode'))
    # Fuera del horario laboral: no se generan alertas ni notificaciones por GPS apagado.
    win = _jornada_window(code)
    if not win or ts < win[0] or ts >= win[1]:
        _exec(db, 'DELETE FROM alerts WHERE code = ?', (code,))
        db.commit()
        return jsonify({'ok': True})
    if body.get('gps'):
        existed = _exec(db, 'SELECT 1 FROM alerts WHERE code = ?', (code,)).fetchone()
        _exec(db, 'DELETE FROM alerts WHERE code = ?', (code,))
        db.commit()
        if existed:
            _send_push_to('ADMIN', name + ' (' + code + ') volvió a activar la ubicación')
    else:
        existed = _exec(db, 'SELECT 1 FROM alerts WHERE code = ?', (code,)).fetchone()
        _exec(db,
              '''INSERT INTO alerts(code, tipo, ts, msj) VALUES (?, 'gps_off', ?, ?)
                 ON CONFLICT(code) DO NOTHING''',
              (code, ts, 'Ubicación apagada o sin señal GPS'))
        db.commit()
        if not existed:
            _send_push_to('ADMIN', '⚠ ' + name + ' (' + code + ') apagó la ubicación. Revisá el panel.')
    return jsonify({'ok': True})


@app.route('/api/alerts')
def alerts_list():
    rows = _exec(get_db(), _q(
        'SELECT code, ts, msj FROM alerts WHERE tipo = ? ORDER BY ts'), ('gps_off',)).fetchall()
    out = []
    for r in rows:
        v = VENDOR_BY_CODE.get(r['code'], {})
        out.append({'code': r['code'], 'name': v.get('name', r['code']),
                    'ts': r['ts'], 'msj': r['msj']})
    return jsonify(out)


@app.route('/api/alerts/visto', methods=['POST'])
def alerts_visto():
    db = get_db()
    _exec(db, 'DELETE FROM alerts WHERE tipo = ?', ('gps_off',))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/start-check', methods=['POST'])
def start_check():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    db = get_db()
    aus = _exec(db, 'SELECT motivo FROM absence WHERE code=? AND fecha=?',
                (code, _today_str())).fetchone()
    if aus:
        return jsonify({'ok': False, 'error': 'Hoy no es tu día de trabajo (%s).' % aus['motivo']}), 403
    try:
        lat = float(body.get('lat'))
        lon = float(body.get('lon'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'coordenadas invalidas'}), 400
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return jsonify({'ok': False, 'error': 'coordenadas fuera de rango'}), 400
    best = None
    for p in _pdv_snapshot():
        pl, po = p.get('lat'), p.get('lon')
        if pl and po and (pl != 0 or po != 0):
            d = _haversine(lat, lon, pl, po)
            if best is None or d < best[0]:
                best = (d, p)
    if not best:
        return jsonify({'ok': False, 'error': 'sin PDV'}), 404
    d, p = best
    return jsonify({
        'ok': True,
        'dist_m': int(d),
        'dentro': d <= max(START_RADIUS_M, _radius_for(p['c'])),
        'pdv': {'cliente': p['c'], 'razon': p['r']}
    })


@app.route('/api/shift')
def shift_get():
    code = request.args.get('code', '').strip().upper()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    row = _exec(get_db(), _q('SELECT shift_ms FROM shifts WHERE code=?'), (code,)).fetchone()
    win = _jornada_window(code)
    now = int(time.time() * 1000)
    return jsonify({
        'ok': True,
        'code': code,
        'shift_ms': row['shift_ms'] if row else SHIFT_DEFAULT_MS,
        'start_ts': win[0] if win else None,
        'end_ts': win[1] if win else None,
        'dentro': bool(win and win[0] <= now < win[1]),
        'jornada_hoy': bool(win),
    })


@app.route('/api/config/shift', methods=['POST'])
def config_shift():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    try:
        hours = float(body.get('shift_hours'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'horas invalidas'}), 400
    if not (1 <= hours <= 24):
        return jsonify({'ok': False, 'error': 'la jornada debe ser entre 1 y 24 horas'}), 400
    shift_ms = int(hours * 3600 * 1000)
    db = get_db()
    _exec(db, '''INSERT INTO shifts(code, shift_ms) VALUES (?,?)
                 ON CONFLICT(code) DO UPDATE SET shift_ms = excluded.shift_ms''', (code, shift_ms))
    db.commit()
    return jsonify({'ok': True, 'shift_ms': shift_ms})


@app.route('/api/absence')
def absence_get():
    code = request.args.get('code', '').strip().upper()
    fecha = request.args.get('fecha', '').strip()
    sql = 'SELECT fecha, code, motivo, ts FROM absence'
    conds, params = [], []
    if code:
        conds.append('code = ?')
        params.append(code)
    if fecha:
        conds.append('fecha = ?')
        params.append(fecha)
    if conds:
        sql += ' WHERE ' + ' AND '.join(conds)
    sql += ' ORDER BY fecha DESC, ts DESC LIMIT 100'
    rows = _exec(get_db(), sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/absence', methods=['POST'])
def absence_post():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    fecha = str(body.get('fecha', '')).strip()
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', fecha):
        return jsonify({'ok': False, 'error': 'fecha invalida'}), 400
    motivo = str(body.get('motivo', '')).strip()[:60] or 'Día libre'
    db = get_db()
    _exec(db, '''INSERT INTO absence(fecha, code, motivo, ts) VALUES (?,?,?,?)
                 ON CONFLICT(fecha, code) DO UPDATE SET motivo = excluded.motivo''',
          (fecha, code, motivo, int(time.time() * 1000)))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/absence', methods=['DELETE'])
def absence_delete():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    fecha = str(body.get('fecha', '')).strip()
    db = get_db()
    _exec(db, 'DELETE FROM absence WHERE code=? AND fecha=?', (code, fecha))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/positions')
def positions():
    db = get_db()
    now = int(time.time() * 1000)
    rows = _exec(db, '''
        SELECT p.code, p.name, p.lat, p.lon, p.ts, p.battery, v.prov, v.color, v.grupo,
               d.app, d.app_version, d.version_code
        FROM positions p
        JOIN (SELECT code, MAX(ts) AS mts FROM positions GROUP BY code) mx
          ON p.code = mx.code AND p.ts = mx.mts
        JOIN vendors v ON v.code = p.code
        LEFT JOIN devices d ON d.code = p.code
    ''').fetchall()
    devs = {d['code']: d for d in _exec(db, 'SELECT code, app, app_version, version_code FROM devices').fetchall()}
    alerts = {a['code']: a['ts'] for a in _exec(db, 'SELECT code, ts FROM alerts').fetchall()}
    out = []
    for r in rows:
        out.append({
            'code': r['code'],
            'name': r['name'],
            'prov': r['prov'],
            'color': r['color'],
            'grupo': r['grupo'],
            'lat': r['lat'],
            'lon': r['lon'],
            'ts': r['ts'],
            'active': (now - r['ts']) < ACTIVE_MS,
            'last': now - r['ts'],
            'gpsAlert': r['code'] in alerts,
            'gpsAlertTs': alerts.get(r['code']),
            'battery': r['battery'],
            'app': r['app'] if r['app'] else 'web',
            'appVersion': r['app_version'] if r['app_version'] else '',
            'appVersionCode': r['version_code'] if r['version_code'] else 0,
        })
    for code, a_ts in alerts.items():
        if code in {r['code'] for r in rows}:
            continue
        v = VENDOR_BY_CODE.get(code)
        if not v:
            continue
        out.append({
            'code': code,
            'name': v.get('name', code),
            'prov': v.get('prov', ''),
            'color': v.get('color', '#666666'),
            'grupo': v.get('grupo', 'rutas'),
            'lat': None,
            'lon': None,
            'ts': None,
            'active': False,
            'last': None,
            'gpsAlert': True,
            'gpsAlertTs': a_ts,
            'battery': None,
            'app': (devs.get(code) or {}).get('app') or 'web',
            'appVersion': (devs.get(code) or {}).get('app_version') or '',
            'appVersionCode': (devs.get(code) or {}).get('version_code') or 0,
        })
    return jsonify(out)


@app.route('/api/history')
def history():
    code = request.args.get('code', '').strip().upper()
    fecha = request.args.get('fecha', '').strip()
    try:
        days = max(1, min(int(request.args.get('days', 1)), 14))
    except (TypeError, ValueError):
        days = 1
    if fecha:
        s, e = _day_range(fecha)
        rows = _exec(get_db(), _q(
            'SELECT lat, lon, ts FROM positions WHERE code = ? AND ts >= ? AND ts < ? ORDER BY ts'),
            (code, s, e)).fetchall()
    else:
        since = int(time.time() * 1000) - days * 86400000
        rows = _exec(get_db(), _q(
            'SELECT lat, lon, ts FROM positions WHERE code = ? AND ts >= ? ORDER BY ts'),
            (code, since)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/vendors')
def vendors_api():
    rows = _exec(get_db(),
                 'SELECT code, name, prov, color, grupo FROM vendors ORDER BY prov, name').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/config/vendor', methods=['POST'])
def config_vendor():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    v = VENDOR_BY_CODE.get(code)
    if not v:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404

    name = body.get('name')
    color = body.get('color')
    if name is not None:
        name = str(name).strip()
        if not name or len(name) > 60:
            return jsonify({'ok': False, 'error': 'nombre invalido'}), 400
        v['name'] = name
    if color is not None:
        color = str(color).strip()
        if not re.fullmatch(r'#[0-9a-fA-F]{6}', color):
            return jsonify({'ok': False, 'error': 'color invalido'}), 400
        v['color'] = color

    OVERRIDES[code] = {'name': v['name'], 'color': v['color']}
    try:
        with open(OVERRIDES_FILE, 'w', encoding='utf-8') as f:
            json.dump(OVERRIDES, f, ensure_ascii=False, indent=2)
    except Exception:
        app.logger.error('Failed to save overrides')

    db = get_db()
    _exec(db, 'UPDATE vendors SET name=?, color=? WHERE code=?', (v['name'], v['color'], code))
    db.commit()
    return jsonify({'ok': True, 'vendor': {'code': code, 'name': v['name'], 'color': v['color']}})


@app.route('/api/pdv')
def pdv_api():
    prov = request.args.get('prov', '').strip().upper()
    q = request.args.get('q', '').strip().lower()
    vta = request.args.get('vta', '').strip().lower()
    items = PDV
    if prov:
        items = [p for p in items if p['prov'] == prov]
    if vta:
        items = [p for p in items if vta in (p.get('vta') or '').lower()]
    if q:
        items = [p for p in items
                 if q in (p.get('r') or '').lower()
                 or q in (p.get('c') or '').lower()
                 or q in (p.get('calle') or '').lower()
                 or q in (p.get('vta') or '').lower()]
        items = items[:50]
    out = []
    for p in items:
        out.append(dict(p, radius=_radius_for(p.get('c', ''))))
    return jsonify(out)


@app.route('/api/pdv', methods=['POST'])
def pdv_post():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    rate_code = code or 'anon'
    if not _rate_limiter.check('pdv:' + rate_code, 5):
        return jsonify({'ok': False, 'error': 'rate limit'}), 429
    v = VENDOR_BY_CODE.get(code)
    if not v:
        return jsonify({'ok': False, 'error': 'codigo de merchan invalido'}), 403
    try:
        lat = float(body.get('lat'))
        lon = float(body.get('lon'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'ubicacion invalida'}), 400
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return jsonify({'ok': False, 'error': 'ubicacion fuera de rango'}), 400
    razon = str(body.get('razon', '')).strip()
    if not razon:
        return jsonify({'ok': False, 'error': 'falta la razon social'}), 400
    app.logger.info('new PDV: %s by %s', razon, code)
    calle = str(body.get('calle', '')).strip()
    altura = str(body.get('altura', '')).strip()
    vta = str(body.get('vta', '')).strip()
    prov = str(body.get('prov', '')).strip().upper()
    if prov not in ('TUCUMAN', 'CATAMARCA'):
        prov = v.get('prov', '')
    telefono = str(body.get('telefono', '')).strip()
    contacto = str(body.get('contacto', '')).strip()
    notas = str(body.get('notas', '')).strip()
    ts = int(time.time() * 1000)
    cliente = 'M' + str(ts)
    db = get_db()
    _exec(db, 'INSERT INTO pdvs_extra(cliente, razon, calle, altura, vta, prov, telefono, contacto, notas, lat, lon, creado_por, ts) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
          (cliente, razon, calle, altura, vta, prov, telefono, contacto, notas, lat, lon, code, ts))
    db.commit()
    nuevo = {'c': cliente, 'r': razon, 'calle': calle, 'altura': altura, 'vta': vta, 'prov': prov,
             'telefono': telefono, 'contacto': contacto, 'notas': notas, 'lat': lat, 'lon': lon}
    with _PDV_LOCK:
        PDV.append(nuevo)
        PDV_BY_CODE[cliente] = nuevo
        PDV_EXTRA.append(nuevo)
    return jsonify({'ok': True, 'cliente': cliente, 'razon': razon, 'lat': lat, 'lon': lon})


@app.route('/api/pdv/radius', methods=['POST'])
def pdv_radius_post():
    body = request.get_json(force=True, silent=True) or {}
    cliente = str(body.get('cliente', '')).strip()
    p = PDV_BY_CODE.get(cliente)
    if not p:
        return jsonify({'ok': False, 'error': 'PDV no encontrado'}), 404
    try:
        radius = int(body.get('radius_m'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'radio invalido'}), 400
    if not (20 <= radius <= 2000):
        return jsonify({'ok': False, 'error': 'el radio debe estar entre 20 y 2000 m'}), 400
    db = get_db()
    _exec(db, '''INSERT INTO pdv_radius(cliente, radius_m) VALUES (?,?)
                 ON CONFLICT(cliente) DO UPDATE SET radius_m = excluded.radius_m''', (cliente, radius))
    db.commit()
    _PDV_RADIUS[cliente] = radius
    return jsonify({'ok': True, 'radius_m': radius})


@app.route('/api/rutas')
def rutas_api():
    merchan = request.args.get('merchan', '').strip().upper()
    dia = request.args.get('dia', '').strip().upper()
    out = []
    for m, dias in RUTAS.items():
        if merchan and m != merchan:
            continue
        for d, items in dias.items():
            if dia and d != dia:
                continue
            for i, (c, razon_fb) in enumerate(items, 1):
                p = PDV_BY_CODE.get(c) or {}
                out.append({
                    'merchan': m,
                    'dia': d,
                    'orden': i,
                    'cliente': c,
                    'razon': p.get('r') or razon_fb,
                    'calle': p.get('calle', ''),
                    'altura': p.get('altura', ''),
                    'vta': p.get('vta', ''),
                    'lat': p.get('lat'),
                    'lon': p.get('lon'),
                    'radius': _radius_for(c),
                })
    return jsonify(out)


@app.route('/api/rutas/version')
def rutas_version():
    merchan = request.args.get('merchan', '').strip().upper()
    h = hashlib.md5()
    for m, dias in RUTAS.items():
        if merchan and m != merchan:
            continue
        h.update(m.encode('utf-8'))
        for d in sorted(dias):
            h.update(d.encode('utf-8'))
            for c, razon_fb in dias[d]:
                h.update(('|%s|%s' % (c, razon_fb)).encode('utf-8'))
    return jsonify({'version': h.hexdigest(), 'fecha': _today_str()})


@app.route('/api/resumen')
def resumen():
    db = get_db()
    fecha = request.args.get('fecha', '').strip()
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', fecha):
        fecha = _today_str()
    key = WEEKDAY_KEYS[datetime.strptime(fecha, '%Y-%m-%d').weekday()]
    s, e = _day_range(fecha)
    now = int(time.time() * 1000)
    out = {}
    for v in VENDORS:
        if v.get('grupo') != 'merchan':
            continue
        code = v['code']
        rows = _exec(db, _q(
            'SELECT lat, lon, ts FROM positions WHERE code = ? AND ts >= ? AND ts < ? ORDER BY ts'),
            (code, s, e)).fetchall()
        km = 0.0
        prev = None
        for r in rows:
            if prev is not None and (r['lat'] != prev[0] or r['lon'] != prev[1]):
                km += _haversine(prev[0], prev[1], r['lat'], r['lon'])
            prev = (r['lat'], r['lon'])
        vis_rows = _exec(db, 'SELECT cliente FROM visitas WHERE code = ? AND fecha = ?',
                         (code, fecha)).fetchall()
        vis = [r['cliente'] for r in vis_rows]
        ruta_codes = [c for c, _ in RUTAS.get(code, {}).get(key, [])]
        vis_ruta = [c for c in vis if c in ruta_codes]
        last_ts = rows[-1]['ts'] if rows else None
        out[code] = {
            'km': round(km / 1000, 2),
            'visitas': len(vis),
            'ruta_total': len(ruta_codes),
            'ruta_vis': len(vis_ruta),
            'last': (now - last_ts) if last_ts else None,
            'active': bool(last_ts and (now - last_ts) < ACTIVE_MS),
        }
    return jsonify(out)


@app.route('/api/jornadas')
def jornadas():
    now = int(time.time() * 1000)
    out = {}
    for v in VENDORS:
        if v.get('grupo') != 'merchan':
            continue
        code = v['code']
        win = _jornada_window(code)
        out[code] = {
            'jornada_hoy': bool(win),
            'start_ts': win[0] if win else None,
            'end_ts': win[1] if win else None,
            'dentro': bool(win and win[0] <= now < win[1]),
        }
    return jsonify(out)


@app.route('/api/nearest-pdv')
def nearest_pdv():
    code = request.args.get('code', '').strip().upper()
    now = int(time.time() * 1000)
    row = _exec(get_db(), _q(
        'SELECT lat, lon, ts FROM positions WHERE code = ? AND ts >= ? ORDER BY ts DESC LIMIT 1'),
        (code, now - ACTIVE_MS)).fetchone()
    if not row:
        return jsonify({'ok': False, 'error': 'sin posicion reciente (el vendedor no inició el envío en vivo)'}), 404
    lat, lon = row['lat'], row['lon']
    best = None
    for p in _pdv_snapshot():
        pl, po = p.get('lat'), p.get('lon')
        if pl and po and (pl != 0 or po != 0):
            d = _haversine(lat, lon, pl, po)
            if best is None or d < best[0]:
                best = (d, p)
    if not best:
        return jsonify({'ok': False, 'error': 'sin pdv'}), 404
    d, p = best
    return jsonify({
        'ok': True,
        'dist_m': int(d),
        'cliente': p['c'],
        'razon': p['r'],
        'calle': p.get('calle', ''),
        'altura': p.get('altura', ''),
        'vta': p.get('vta', ''),
        'lat': p['lat'],
        'lon': p['lon'],
        'radius': _radius_for(p['c']),
        'ts': row['ts'],
    })


@app.route('/api/visitas', methods=['GET'])
def visitas_get():
    code = request.args.get('code', '').strip().upper()
    fecha = request.args.get('fecha', '').strip()
    sql = 'SELECT fecha, code, cliente, razon, calle, vta, ts, foto_ts, foto_salida_ts FROM visitas'
    conds, params = [], []
    if code:
        conds.append('code = ?')
        params.append(code)
    if fecha:
        conds.append('fecha = ?')
        params.append(fecha)
    if conds:
        sql += ' WHERE ' + ' AND '.join(conds)
    sql += ' ORDER BY ts'
    rows = _exec(get_db(), sql, params).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d['has_foto'] = 1 if (d.get('foto_ts') and d['foto_ts'] > 0) else 0
        d['finalizada'] = 1 if (d.get('foto_salida_ts') and d['foto_salida_ts'] > 0) else 0
        d.pop('foto_ts', None)
        d.pop('foto_salida_ts', None)
        out.append(d)
    return jsonify(out)


@app.route('/api/visitas/foto')
def visitas_foto():
    code = request.args.get('code', '').strip().upper()
    cliente = request.args.get('cliente', '').strip()
    fecha = request.args.get('fecha', '').strip() or _today_str()
    row = _exec(get_db(), _q(
        'SELECT foto, foto_ts FROM visitas WHERE code=? AND cliente=? AND fecha=?'),
        (code, cliente, fecha)).fetchone()
    if not row or not row['foto']:
        return jsonify({'ok': False, 'error': 'sin foto'}), 404
    return jsonify({'ok': True, 'foto': row['foto'], 'foto_ts': row['foto_ts']})


@app.route('/api/visitas/foto/img')
def visitas_foto_img():
    code = request.args.get('code', '').strip().upper()
    cliente = request.args.get('cliente', '').strip()
    fecha = request.args.get('fecha', '').strip() or _today_str()
    row = _exec(get_db(), _q(
        'SELECT foto FROM visitas WHERE code=? AND cliente=? AND fecha=?'),
        (code, cliente, fecha)).fetchone()
    if not row or not row['foto']:
        return 'sin foto', 404
    data = base64.b64decode(row['foto'])
    if data[:2] == b'\xff\xd8':
        ctype = 'image/jpeg'
    elif data[:4] == b'\x89PNG':
        ctype = 'image/png'
    elif data[:3] == b'GIF':
        ctype = 'image/gif'
    elif data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        ctype = 'image/webp'
    else:
        ctype = 'image/jpeg'
    return Response(data, mimetype=ctype)


@app.route('/api/visitas/resumen')
def visitas_resumen():
    fecha = request.args.get('fecha', '').strip() or _today_str()
    rows = _exec(get_db(),
                 'SELECT code, cliente, foto_ts FROM visitas WHERE fecha = ? ORDER BY ts', (fecha,)).fetchall()
    out = {}
    for r in rows:
        item = out.setdefault(r['code'], {'n': 0, 'clientes': [], 'fotos': {}})
        item['n'] += 1
        item['clientes'].append(r['cliente'])
        item['fotos'][r['cliente']] = 1 if (r['foto_ts'] and r['foto_ts'] > 0) else 0
    return jsonify(out)


@app.route('/api/merchan-pdv')
def merchan_pdv():
    now = int(time.time() * 1000)
    db = get_db()
    rows = _exec(db, '''
        SELECT p.code, p.lat, p.lon FROM positions p
        JOIN (SELECT code, MAX(ts) AS mts FROM positions GROUP BY code) mx
          ON p.code = mx.code AND p.ts = mx.mts
        WHERE p.ts >= ?
    ''', (now - ACTIVE_MS,)).fetchall()
    dia = _today_key()
    fecha = _today_str()
    out = {}
    for r in rows:
        lat, lon = r['lat'], r['lon']
        best = None
        for p in _pdv_snapshot():
            pl, po = p.get('lat'), p.get('lon')
            if pl and po and (pl != 0 or po != 0):
                d = _haversine(lat, lon, pl, po)
                if best is None or d < best[0]:
                    best = (d, p)
        if best:
            d, p = best
            route_set = {c for c, _ in RUTAS.get(r['code'], {}).get(dia, [])}
            vis = set(row['cliente'] for row in _exec(
                db, 'SELECT cliente FROM visitas WHERE code=? AND fecha=?', (r['code'], fecha)).fetchall())
            cliente = p['c']
            out[r['code']] = {
                'cliente': cliente,
                'razon': p['r'],
                'calle': p.get('calle', ''),
                'altura': p.get('altura', ''),
                'dist_m': int(d),
                'radius': _radius_for(cliente),
                'dentro': d <= _radius_for(cliente),
                'pendiente': cliente in route_set and cliente not in vis,
                'lat': p['lat'],
                'lon': p['lon'],
            }
    return jsonify(out)


@app.route('/api/geoevents')
def geoevents():
    code = request.args.get('code', '').strip().upper()
    try:
        limit = max(1, min(int(request.args.get('limit', 20)), 200))
    except (TypeError, ValueError):
        limit = 20
    db = get_db()
    if code:
        rows = _exec(db, 'SELECT code, cliente, razon, tipo, dist_m, ts FROM geoevents WHERE code=? ORDER BY ts DESC LIMIT ?',
                     (code, limit)).fetchall()
    else:
        rows = _exec(db, 'SELECT code, cliente, razon, tipo, dist_m, ts FROM geoevents ORDER BY ts DESC LIMIT ?',
                     (limit,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/visitas', methods=['POST'])
def visitas_post():
    foto = None
    foto_ts = None
    if request.content_type and 'multipart' in request.content_type:
        body = {k: request.form.get(k, '') for k in request.form}
        f = request.files.get('foto')
        if f and f.filename:
            foto = base64.b64encode(f.read()).decode('ascii')
            foto_ts = int(time.time() * 1000)
    else:
        body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    rate_code = code or 'anon'
    if not _rate_limiter.check('visita:' + rate_code, 10):
        return jsonify({'ok': False, 'error': 'rate limit'}), 429
    v = VENDOR_BY_CODE.get(code)
    if not v:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 403
    cliente = str(body.get('cliente', '')).strip()
    if not cliente:
        return jsonify({'ok': False, 'error': 'falta cliente'}), 400
    app.logger.info('visita: %s -> %s', code, cliente)
    fecha = str(body.get('fecha', '')).strip() or _today_str()
    if len(fecha) != 10:
        return jsonify({'ok': False, 'error': 'fecha invalida'}), 400
    pdv = PDV_BY_CODE.get(cliente)
    razon = pdv.get('r', '') if pdv else str(body.get('razon', '')).strip()
    calle = pdv.get('calle', '') if pdv else str(body.get('calle', '')).strip()
    altura = pdv.get('altura', '') if pdv else str(body.get('altura', '')).strip()
    vta = pdv.get('vta', '') if pdv else str(body.get('vta', '')).strip()
    if not pdv and not razon:
        return jsonify({'ok': False, 'error': 'cliente no encontrado en PDV'}), 404
    db = get_db()
    if body.get('validate'):
        row = _exec(db, 'SELECT lat, lon, ts FROM positions WHERE code = ? ORDER BY ts DESC LIMIT 1',
                    (code,)).fetchone()
        now = int(time.time() * 1000)
        c_lat = c_lon = None
        try:
            if body.get('lat') not in (None, ''):
                c_lat = float(body.get('lat'))
            if body.get('lon') not in (None, ''):
                c_lon = float(body.get('lon'))
        except (TypeError, ValueError):
            c_lat = c_lon = None
        if c_lat is not None and c_lon is not None and row and now - row['ts'] <= VISIT_MAX_MS:
            if _haversine(c_lat, c_lon, row['lat'], row['lon']) > 1000:
                return jsonify({'ok': False, 'error': 'La posición del celular no coincide con la enviada en vivo. Reintentá en unos segundos.'}), 400
        if c_lat is not None and c_lon is not None:
            use_lat, use_lon = c_lat, c_lon
        elif row and now - row['ts'] <= VISIT_MAX_MS:
            use_lat, use_lon = row['lat'], row['lon']
        else:
            use_lat = use_lon = None
        if use_lat is None:
            return jsonify({'ok': False, 'error': 'No hay una posición para validar. Activá el GPS y acercate al PDV.'}), 400
        if pdv and pdv.get('lat') and pdv.get('lon') and (pdv['lat'] != 0 or pdv['lon'] != 0):
            try:
                acc = float(body.get('accuracy') or 0)
            except (TypeError, ValueError):
                acc = 0
            limit = _radius_for(cliente) + max(0, acc - 50)  # tolerancia al ruido del GPS
            d = _haversine(use_lat, use_lon, pdv['lat'], pdv['lon'])
            if d > limit:
                return jsonify({'ok': False, 'error': 'No estás en el PDV (%s). Estás a %d m de ese PDV.' % (pdv.get('r', ''), int(d))}), 400
    # PDV de ruta que no estaba georreferenciado: al primer registro de visita
    # con GPS, queda cargado en el catálogo (pdvs_extra) con el código real.
    if pdv is None and razon:
        g_lat = g_lon = None
        try:
            if body.get('lat') not in (None, ''):
                g_lat = float(body.get('lat'))
            if body.get('lon') not in (None, ''):
                g_lon = float(body.get('lon'))
        except (TypeError, ValueError):
            g_lat = g_lon = None
        if g_lat is not None and g_lon is not None and -90 <= g_lat <= 90 and -180 <= g_lon <= 180:
            nprov = v.get('prov', '')
            _exec(db, '''INSERT INTO pdvs_extra(cliente, razon, calle, altura, vta, prov, telefono, contacto, notas, lat, lon, creado_por, ts)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                         ON CONFLICT(cliente) DO UPDATE SET
                           razon=excluded.razon, calle=excluded.calle, altura=excluded.altura, vta=excluded.vta,
                           prov=excluded.prov, lat=excluded.lat, lon=excluded.lon, ts=excluded.ts''',
                  (cliente, razon, calle, altura, vta, nprov, '', '', '', g_lat, g_lon, code, int(time.time() * 1000)))
            db.commit()
            nuevo = {'c': cliente, 'r': razon, 'calle': calle, 'altura': altura, 'vta': vta,
                     'prov': nprov, 'lat': g_lat, 'lon': g_lon}
            if not PDV_BY_CODE.get(cliente):
                with _PDV_LOCK:
                    PDV.append(nuevo)
            PDV_BY_CODE[cliente] = nuevo
            if not any(p.get('c') == cliente for p in PDV_EXTRA):
                with _PDV_LOCK:
                    PDV_EXTRA.append(nuevo)
    ts = int(time.time() * 1000)
    if PG:
        _exec(db, '''INSERT INTO visitas(fecha, code, cliente, razon, calle, vta, ts, foto, foto_ts, foto_salida_ts)
                     VALUES (?,?,?,?,?,?,?,?,?,NULL)
                     ON CONFLICT(fecha, code, cliente) DO UPDATE SET
                       razon=excluded.razon, calle=excluded.calle, vta=excluded.vta,
                       ts=excluded.ts, foto=excluded.foto, foto_ts=excluded.foto_ts, foto_salida_ts=NULL''',
              (fecha, code, cliente, razon, calle, vta, ts, foto, foto_ts))
    else:
        _exec(db, 'INSERT OR REPLACE INTO visitas(fecha, code, cliente, razon, calle, vta, ts, foto, foto_ts, foto_salida_ts) VALUES (?,?,?,?,?,?,?,?,?,NULL)',
              (fecha, code, cliente, razon, calle, vta, ts, foto, foto_ts))
    db.commit()
    try:
        admin_subs = _exec(db, "SELECT endpoint, p256dh, auth FROM push_subs WHERE code = 'ADMIN'").fetchall()
        if admin_subs:
            vendor_name = v.get('name', code) if v else code
            payload = json.dumps({'title': 'Visita registrada', 'body': vendor_name + ' → ' + str(razon)})
            for s in admin_subs:
                _send_push_to(s['endpoint'], s['p256dh'], s['auth'], payload)
    except Exception:
        pass
    return jsonify({'ok': True, 'ts': ts, 'has_foto': 1 if foto else 0})


@app.route('/api/visitas', methods=['DELETE'])
def visitas_delete():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    cliente = str(body.get('cliente', '')).strip()
    fecha = str(body.get('fecha', '')).strip() or _today_str()
    db = get_db()
    _exec(db, 'DELETE FROM visitas WHERE code = ? AND cliente = ? AND fecha = ?',
          (code, cliente, fecha))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/visitas/finalizar', methods=['POST'])
def visitas_finalizar():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    cliente = str(body.get('cliente', '')).strip()
    fecha = str(body.get('fecha', '')).strip() or _today_str()
    v = VENDOR_BY_CODE.get(code)
    if not v:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 403
    if not cliente:
        return jsonify({'ok': False, 'error': 'falta cliente'}), 400
    db = get_db()
    row = _exec(db, 'SELECT cliente FROM visitas WHERE fecha = ? AND code = ? AND cliente = ?',
                (fecha, code, cliente)).fetchone()
    if not row:
        return jsonify({'ok': False, 'error': 'No hay visita registrada para ese PDV hoy.'}), 404
    ts_salida = int(time.time() * 1000)
    _exec(db, 'UPDATE visitas SET foto_salida_ts = ? WHERE fecha = ? AND code = ? AND cliente = ?',
          (ts_salida, fecha, code, cliente))
    db.commit()
    return jsonify({'ok': True, 'foto_salida_ts': ts_salida})


# ---------------- Mensajes panel -> merchan ----------------

@app.route('/api/messages')
def messages_get():
    code = request.args.get('code', '').strip().upper()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    rows = _exec(get_db(),
                 'SELECT id, code, msj, ts, visto FROM messages WHERE code=? AND visto=0 ORDER BY ts DESC LIMIT 20',
                 (code,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/messages/visto', methods=['POST'])
def messages_visto():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    db = get_db()
    _exec(db, 'UPDATE messages SET visto=1 WHERE code=? AND visto=0', (code,))
    db.commit()
    return jsonify({'ok': True})


def _store_message(db, code, msj, now):
    _exec(db, 'INSERT INTO messages(code, msj, ts, visto) VALUES (?,?,?,0)', (code, msj, now))


def _send_push_to(code, msj):
    try:
        from pywebpush import webpush
    except Exception:
        logging.getLogger(__name__).error('Failed to import pywebpush')
        return
    keys = _vapid_keys()
    db = get_db()
    subs = _exec(db, 'SELECT endpoint, p256dh, auth FROM push_subs WHERE code=?', (code,)).fetchall()
    for s in subs:
        try:
            webpush(
                subscription_info={'endpoint': s['endpoint'],
                                   'keys': {'p256dh': s['p256dh'], 'auth': s['auth']}},
                data=json.dumps({'title': 'GPS Merchan', 'body': msj}),
                vapid_private_key=keys['priv'],
                vapid_claims={'sub': VAPID_MAILTO},
            )
        except Exception:
            logging.getLogger(__name__).error('Failed to send push notification')


@app.route('/api/comunicar', methods=['POST'])
def comunicar():
    body = request.get_json(force=True, silent=True) or {}
    msj = str(body.get('msj', '')).strip()
    if not msj:
        return jsonify({'ok': False, 'error': 'escribí un mensaje'}), 400
    if len(msj) > 500:
        return jsonify({'ok': False, 'error': 'el mensaje no puede superar 500 caracteres'}), 400
    target = str(body.get('code', '')).strip().upper()
    if target == 'ALL':
        codes = [v['code'] for v in VENDORS if v.get('grupo') == 'merchan']
    elif target in VENDOR_BY_CODE:
        codes = [target]
    else:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    db = get_db()
    now = int(time.time() * 1000)
    for c in codes:
        _store_message(db, c, msj, now)
    db.commit()
    for c in codes:
        _send_push_to(c, msj)
    return jsonify({'ok': True, 'n': len(codes)})


# ---------------- Push (VAPID) ----------------

def _vapid_keys():
    db = get_db()
    row = _exec(db, 'SELECT value FROM settings WHERE key=?', ('vapid_priv',)).fetchone()
    if row:
        pub = _exec(db, 'SELECT value FROM settings WHERE key=?', ('vapid_pub',)).fetchone()
        return {'priv': row['value'], 'pub': pub['value'] if pub else ''}
    from py_vapid import Vapid
    from cryptography.hazmat.primitives.serialization import Encoding, PrivateFormat, PublicFormat, NoEncryption
    v = Vapid()
    v.generate_keys()
    priv_pem = v.private_key.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption()).decode()
    pub_raw = v.public_key.public_bytes(Encoding.X962, PublicFormat.UncompressedPoint)
    pub_b64 = base64.urlsafe_b64encode(pub_raw).decode().rstrip('=')
    _exec(db, '''INSERT INTO settings(key, value) VALUES ('vapid_priv', ?)
                 ON CONFLICT(key) DO UPDATE SET value = excluded.value''', (priv_pem,))
    _exec(db, '''INSERT INTO settings(key, value) VALUES ('vapid_pub', ?)
                 ON CONFLICT(key) DO UPDATE SET value = excluded.value''', (pub_b64,))
    db.commit()
    return {'priv': priv_pem, 'pub': pub_b64}


@app.route('/api/push/vapid')
def push_vapid():
    return jsonify({'publicKey': _vapid_keys()['pub']})


@app.route('/api/push/subscribe', methods=['POST'])
def push_subscribe():
    body = request.get_json(force=True, silent=True) or {}
    code = str(body.get('code', '')).strip().upper()
    if code not in VENDOR_BY_CODE and code != 'ADMIN':
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    sub = body.get('subscription') or {}
    endpoint = str(sub.get('endpoint', '')).strip()
    keys = sub.get('keys') or {}
    p256dh = str(keys.get('p256dh', '')).strip()
    auth = str(keys.get('auth', '')).strip()
    if not endpoint or not p256dh or not auth:
        return jsonify({'ok': False, 'error': 'suscripción invalida'}), 400
    db = get_db()
    _exec(db, '''INSERT INTO push_subs(endpoint, code, p256dh, auth, ts) VALUES (?,?,?,?,?)
                 ON CONFLICT(endpoint) DO UPDATE SET code = excluded.code, p256dh = excluded.p256dh,
                   auth = excluded.auth, ts = excluded.ts''',
          (endpoint, code, p256dh, auth, int(time.time() * 1000)))
    db.commit()
    return jsonify({'ok': True})


# ---------------- Reportes y export ----------------

def _report_code(db, code, desde, hasta):
    d0 = datetime.strptime(desde, '%Y-%m-%d').date()
    d1 = datetime.strptime(hasta, '%Y-%m-%d').date()
    days = []
    tot_sched = tot_vis_ruta = tot_vis_all = 0
    tot_km = tot_h = 0.0
    tot_inicio = tot_fin = None
    dia = d0
    while dia <= d1:
        fecha = dia.isoformat()
        dia_key = WEEKDAY_KEYS[dia.weekday()]
        sched = RUTAS.get(code, {}).get(dia_key, [])
        sched_set = {c for c, _ in sched}
        vis_rows = _exec(db, 'SELECT cliente FROM visitas WHERE code=? AND fecha=?',
                         (code, fecha)).fetchall()
        vis_set = {r['cliente'] for r in vis_rows}
        vis_ruta = len(sched_set & vis_set)
        vis_all = len(vis_set)
        s, e = _day_range(fecha)
        pos = _exec(db, 'SELECT lat, lon, ts FROM positions WHERE code=? AND ts>=? AND ts<? ORDER BY ts',
                    (code, s, e)).fetchall()
        km = 0.0
        act_ms = 0
        prev = None
        for r in pos:
            if prev is not None:
                gap = r['ts'] - prev['ts']
                if 0 <= gap <= ACTIVE_GAP_MS:
                    km += _haversine(prev['lat'], prev['lon'], r['lat'], r['lon'])
                    act_ms += gap
            prev = r
        if pos:
            if tot_inicio is None or pos[0]['ts'] < tot_inicio:
                tot_inicio = pos[0]['ts']
            if tot_fin is None or pos[-1]['ts'] > tot_fin:
                tot_fin = pos[-1]['ts']
        pct = round(vis_ruta * 100.0 / len(sched)) if sched else (100 if vis_all else 0)
        days.append({
            'fecha': fecha,
            'dia': RUTAS_DIA_NAMES[dia_key],
            'scheduled': len(sched),
            'visitados_ruta': vis_ruta,
            'visitas_totales': vis_all,
            'pct': pct,
            'km': round(km / 1000, 1),
            'horas': round(act_ms / 3600000.0, 2),
            'inicio': pos[0]['ts'] if pos else None,
            'fin': pos[-1]['ts'] if pos else None,
        })
        tot_sched += len(sched)
        tot_vis_ruta += vis_ruta
        tot_vis_all += vis_all
        tot_km += km
        tot_h += act_ms
        dia += timedelta(days=1)
    return {
        'code': code,
        'nombre': VENDOR_BY_CODE.get(code, {}).get('name', code),
        'desde': desde,
        'hasta': hasta,
        'days': days,
        'totals': {
            'scheduled': tot_sched,
            'visitados_ruta': tot_vis_ruta,
            'visitas_totales': tot_vis_all,
            'pct': round(tot_vis_ruta * 100.0 / tot_sched) if tot_sched else 0,
            'km': round(tot_km / 1000, 1),
            'horas': round(tot_h / 3600000.0, 2),
            'inicio': tot_inicio,
            'fin': tot_fin,
        },
    }


def _estadias_dia(db, code, fecha):
    dia_key = WEEKDAY_KEYS[datetime.strptime(fecha, '%Y-%m-%d').date().weekday()]
    s, e = _day_range(fecha)
    pos = _exec(db, 'SELECT lat, lon, ts FROM positions WHERE code=? AND ts>=? AND ts<? ORDER BY ts',
                (code, s, e)).fetchall()
    if not pos:
        return []
    candidatos = {}
    for c, _ in RUTAS.get(code, {}).get(dia_key, []):
        p = PDV_BY_CODE.get(c)
        if p and p.get('lat') and p.get('lon'):
            candidatos[c] = p
    for r in _exec(db, 'SELECT DISTINCT cliente FROM visitas WHERE code=? AND fecha=?', (code, fecha)).fetchall():
        p = PDV_BY_CODE.get(r['cliente'])
        if p and p.get('lat') and p.get('lon'):
            candidatos[r['cliente']] = p
    if not candidatos:
        return []
    MERGE_GAP_MS = 3 * 60 * 1000
    MIN_SEG_MS = 60 * 1000
    out = []
    for c, p in candidatos.items():
        rad = _radius_for(c)
        seg_inicio = None
        seg_fin = None
        segs = []
        for r in pos:
            d = _haversine(r['lat'], r['lon'], p['lat'], p['lon'])
            if d <= rad:
                if seg_inicio is None:
                    seg_inicio = r['ts']
                seg_fin = r['ts']
            elif seg_inicio is not None and r['ts'] - seg_fin > MERGE_GAP_MS:
                if seg_fin - seg_inicio >= MIN_SEG_MS:
                    segs.append((seg_inicio, seg_fin))
                seg_inicio = None
                seg_fin = None
        if seg_inicio is not None and seg_fin - seg_inicio >= MIN_SEG_MS:
            segs.append((seg_inicio, seg_fin))
        if not segs:
            continue
        vis = _exec(db, 'SELECT COUNT(*) AS n FROM visitas WHERE code=? AND fecha=? AND cliente=?',
                    (code, fecha, c)).fetchone()
        out.append({
            'cliente': c,
            'razon': p.get('r', ''),
            'calle': p.get('calle', ''),
            'vta': p.get('vta', ''),
            'fecha': fecha,
            'dia': RUTAS_DIA_NAMES[dia_key],
            'minutos': round(sum(b - a for a, b in segs) / 60000.0, 1),
            'entradas': len(segs),
            'inicio': segs[0][0],
            'fin': segs[-1][1],
            'visitas': int(vis['n']) if vis else 0,
        })
    out.sort(key=lambda x: x['inicio'])
    return out


def _estadias_dias(db, code, desde, hasta):
    d0 = datetime.strptime(desde, '%Y-%m-%d').date()
    d1 = datetime.strptime(hasta, '%Y-%m-%d').date()
    fechas = set()
    for r in _exec(db, 'SELECT DISTINCT fecha FROM visitas WHERE code=? AND fecha>=? AND fecha<=?',
                   (code, desde, hasta)):
        fechas.add(r['fecha'])
    dia = d0
    while dia <= d1:
        s, e = _day_range(dia.isoformat())
        has = _exec(db, 'SELECT 1 AS x FROM positions WHERE code=? AND ts>=? AND ts<? LIMIT 1',
                    (code, s, e)).fetchone()
        if has:
            fechas.add(dia.isoformat())
        dia += timedelta(days=1)
    for f in sorted(fechas):
        filas = _estadias_dia(db, code, f)
        if filas:
            yield f, filas


@app.route('/api/estadias')
def estadias():
    code = request.args.get('code', '').strip().upper()
    desde = request.args.get('desde', '').strip()
    hasta = request.args.get('hasta', '').strip()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    if not (re.fullmatch(r'\d{4}-\d{2}-\d{2}', desde or '') and re.fullmatch(r'\d{4}-\d{2}-\d{2}', hasta or '')):
        return jsonify({'ok': False, 'error': 'rango de fechas invalido'}), 400
    db = get_db()
    dias = []
    total = 0.0
    for f, filas in _estadias_dias(db, code, desde, hasta):
        total += sum(x['minutos'] for x in filas)
        dias.append({'fecha': f, 'filas': filas})
    return jsonify({'ok': True, 'code': code, 'nombre': VENDOR_BY_CODE[code].get('name', code),
                    'desde': desde, 'hasta': hasta, 'dias': dias,
                    'totales': {'minutos': round(total, 1), 'horas': round(total / 60.0, 2)}})


@app.route('/api/export/estadias.xlsx')
def export_estadias_xlsx():
    code = request.args.get('code', '').strip().upper()
    desde = request.args.get('desde', '').strip()
    hasta = request.args.get('hasta', '').strip()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    if not (re.fullmatch(r'\d{4}-\d{2}-\d{2}', desde or '') and re.fullmatch(r'\d{4}-\d{2}-\d{2}', hasta or '')):
        return jsonify({'ok': False, 'error': 'rango de fechas invalido'}), 400
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        return jsonify({'ok': False, 'error': 'el generador de Excel no está instalado'}), 500
    db = get_db()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tiempo por PDV'
    headers = ['Merchan', 'Codigo', 'Fecha', 'Dia', 'Cliente', 'Razon social', 'Calle', 'Ruta',
               'Entrada', 'Salida', 'Minutos', 'Veces', 'Visitas registradas']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E5E7EB')
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    total_min = 0.0
    for f, filas in _estadias_dias(db, code, desde, hasta):
        for x in filas:
            total_min += x['minutos']
            ws.append([
                VENDOR_BY_CODE[code].get('name', code), code, f, x['dia'], x['cliente'],
                x['razon'], x['calle'], x['vta'] or '',
                (datetime.fromtimestamp(x['inicio'] / 1000, timezone.utc) - timedelta(hours=3)).strftime('%H:%M:%S'),
                (datetime.fromtimestamp(x['fin'] / 1000, timezone.utc) - timedelta(hours=3)).strftime('%H:%M:%S'),
                x['minutos'], x['entradas'], x['visitas'],
            ])
    if total_min > 0:
        ws.append(['TOTAL', '', '', '', '', '', '', '', '', '', round(total_min, 1), '', ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    widths = [18, 8, 11, 10, 10, 34, 24, 12, 10, 10, 9, 8, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r'[^A-Za-z0-9_-]', '_', code)
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=estadias_%s_%s_%s.xlsx' % (safe, desde, hasta)})


@app.route('/api/reporte')
def reporte():
    code = request.args.get('code', '').strip().upper()
    desde = request.args.get('desde', '').strip()
    hasta = request.args.get('hasta', '').strip()
    if code not in VENDOR_BY_CODE:
        return jsonify({'ok': False, 'error': 'codigo invalido'}), 404
    if not (re.fullmatch(r'\d{4}-\d{2}-\d{2}', desde or '') and re.fullmatch(r'\d{4}-\d{2}-\d{2}', hasta or '')):
        return jsonify({'ok': False, 'error': 'rango de fechas invalido'}), 400
    data = _report_code(get_db(), code, desde, hasta)
    data['ok'] = True
    return jsonify(data)


@app.route('/api/export/visitas.xlsx')
def export_visitas_xlsx():
    desde = request.args.get('desde', '').strip()
    hasta = request.args.get('hasta', '').strip()
    code = request.args.get('code', '').strip().upper()
    if desde and not re.fullmatch(r'\d{4}-\d{2}-\d{2}', desde):
        desde = ''
    if hasta and not re.fullmatch(r'\d{4}-\d{2}-\d{2}', hasta):
        hasta = ''
    db = get_db()
    sql = 'SELECT fecha, code, cliente, razon, calle, vta, ts, foto_ts, foto_salida_ts FROM visitas'
    conds, params = [], []
    if code:
        conds.append('code = ?')
        params.append(code)
    if desde:
        conds.append('fecha >= ?')
        params.append(desde)
    if hasta:
        conds.append('fecha <= ?')
        params.append(hasta)
    if conds:
        sql += ' WHERE ' + ' AND '.join(conds)
    sql += ' ORDER BY fecha, code, ts'
    rows = _exec(db, sql, params).fetchall()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        return jsonify({'ok': False, 'error': 'el generador de Excel no está instalado'}), 500
    wb = Workbook()
    ws = wb.active
    ws.title = 'Visitas'
    headers = ['Fecha', 'Codigo', 'Colaborador', 'Cliente', 'Razon social', 'Calle', 'Ruta',
               'Hora entrada', 'Hora salida', 'Con foto']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E5E7EB')
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    for r in rows:
        v = VENDOR_BY_CODE.get(r['code'])
        ts = r['ts']
        hora = (datetime.fromtimestamp(ts / 1000, timezone.utc) - timedelta(hours=3)).strftime('%H:%M:%S') if ts else ''
        hora_salida = ''
        if r['foto_salida_ts'] and r['foto_salida_ts'] > 0:
            hora_salida = (datetime.fromtimestamp(r['foto_salida_ts'] / 1000, timezone.utc) - timedelta(hours=3)).strftime('%H:%M:%S')
        cf = 'Si' if (r['foto_ts'] and r['foto_ts'] > 0) else 'No'
        ws.append([r['fecha'], r['code'], (v.get('name', '') if v else ''), r['cliente'],
                    r['razon'], r['calle'], r['vta'] or '', hora, hora_salida, cf])
    widths = [12, 8, 20, 10, 34, 24, 12, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r'[^A-Za-z0-9_-]', '_', code or 'todos')
    fname = 'visitas_%s_%s_%s.xlsx' % (safe, desde or 'inicio', hasta or 'fin')
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=' + fname})


@app.route('/api/export/recorrido.xlsx')
def export_recorrido_xlsx():
    code = request.args.get('code', '').strip().upper()
    fecha = request.args.get('fecha', '').strip()
    if not code or not fecha or not re.fullmatch(r'\d{4}-\d{2}-\d{2}', fecha):
        return jsonify({'ok': False, 'error': 'code y fecha requeridos'}), 400
    s, e = _day_range(fecha)
    rows = _exec(get_db(),
                 'SELECT lat, lon, ts FROM positions WHERE code=? AND ts>=? AND ts<? ORDER BY ts',
                 (code, s, e)).fetchall()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        return jsonify({'ok': False, 'error': 'el generador de Excel no está instalado'}), 500
    wb = Workbook()
    ws = wb.active
    ws.title = 'Recorrido'
    headers = ['Fecha', 'Hora', 'Latitud', 'Longitud']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E5E7EB')
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    for r in rows:
        dt = datetime.fromtimestamp(r['ts'] / 1000, timezone.utc) - timedelta(hours=3)
        ws.append([fecha, dt.strftime('%H:%M:%S'), r['lat'], r['lon']])
    widths = [12, 10, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r'[^A-Za-z0-9_-]', '_', code)
    fname = 'recorrido_%s_%s.xlsx' % (safe, fecha)
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=' + fname})


# ---------------- Páginas ----------------

@app.route('/tracker/<code>')
def tracker(code):
    vendor = VENDOR_BY_CODE.get(code.strip().upper())
    if not vendor:
        return '<h3>Código inválido</h3><p>Revisá el enlace con tu coordinador.</p>', 404
    return render_template('tracker.html', vendor=vendor)


@app.route('/mapa-pdv')
def mapa_pdv():
    prov = request.args.get('prov', '').strip().upper()
    return Response(
        render_template('mapa_pdv.html',
                        prov=prov if prov in ('TUCUMAN', 'CATAMARCA') else ''),
        headers={'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0'})


@app.route('/mapa-pdv-tucuman')
def mapa_pdv_tucuman():
    return redirect('/mapa-pdv?prov=TUCUMAN')


@app.route('/mapa-pdv-catamarca')
def mapa_pdv_catamarca():
    return redirect('/mapa-pdv?prov=CATAMARCA')


@app.route('/', methods=['GET', 'POST'])
def dashboard():
    if request.method == 'POST':
        if request.form.get('pin') == DASH_PIN:
            session['auth'] = True
        return redirect('/')
    if not session.get('auth'):
        return render_template('dashboard.html', locked=True, wrong=request.args.get('w') == '1')
    vendors = _exec(get_db(),
                    'SELECT code, name, prov, color, grupo FROM vendors ORDER BY prov, name').fetchall()
    return render_template('dashboard.html', locked=False, vendors=[dict(v) for v in vendors])


@app.route('/logout')
def logout():
    session.pop('auth', None)
    return redirect('/')


@app.route('/manifest.json')
def manifest():
    code = request.args.get('code', '').strip().upper()
    if code in VENDOR_BY_CODE:
        start_url = '/tracker/' + code
    else:
        start_url = '/tracker/'
    return jsonify({
        'name': 'GPS Merchan',
        'short_name': 'GPS Merchan',
        'start_url': start_url,
        'scope': '/',
        'display': 'standalone',
        'background_color': '#0f1420',
        'theme_color': '#0f1420',
        'icons': [
            {'src': '/static/icon-192.png', 'sizes': '192x192', 'type': 'image/png'},
            {'src': '/static/icon-512.png', 'sizes': '512x512', 'type': 'image/png'},
            {'src': '/static/icon-512.png', 'sizes': '512x512', 'type': 'image/png', 'purpose': 'maskable'},
        ],
    })


@app.route('/sw.js')
def sw():
    with open(os.path.join(BASE_DIR, 'sw.js'), encoding='utf-8') as f:
        return (f.read(), 200, {'Content-Type': 'application/javascript; charset=utf-8', 'Service-Worker-Allowed': '/'})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, threaded=True, debug=True)
