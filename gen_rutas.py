# -*- coding: utf-8 -*-
"""Genera rutas.json con la asignación semanal merchan -> dia -> [codigos PDV].

Lee los Excel de la carpeta ruteos (hojas de día LUNES..SÁBADO, ignora
'Faltantes' y 'Semana'), extrae los códigos de PDV y los guarda por merchan.
Cada archivo se identifica por su nombre (contiene el apellido del merchan).
"""
import json
import os
import shutil
import sys
import tempfile
import unicodedata

import openpyxl

RUTEOS_DIR = r'C:\Users\gborrasar\Desktop\ruteos'
PDV_FILE = r'C:\Users\gborrasar\Documents\Default Project\gps-vendedores\pdv.json'
OUT = r'C:\Users\gborrasar\Documents\Default Project\gps-vendedores\rutas.json'

# nombre exacto de archivo -> codigo de merchan
FILE_MAP = {
    'AUGUSTO MADRID.xlsx': 'MADRID-38',
    'carlos lazo.xlsx': 'LAZO-33',
    'Corbalan Facundo.xlsx': 'CORBALAN-31',
    'GONZALO AGUILAR 18-07-26 a 24-07- 2026).xlsx': 'AGUILAR-32',
    'Julian Albornoz.xlsx': 'ALBONOZ-39',
    'Leonardo Silva.xlsx': 'SILVA-37',
    'Matias Abib.xlsx': 'ABIB-34',
    'Matias Emeterio.xlsx': 'EMETERIO-35',
    'Santiago Vera .xlsx': 'VERA-36',
}

DIAS = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']


def norm(s):
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().upper().strip()


def day_of_sheet(name):
    n = norm(name)
    for d in DIAS:
        if d in n:
            return d
    return None


def load_pdv():
    with open(PDV_FILE, encoding='utf-8') as f:
        return json.load(f)


def extract_codes(path):
    """Devuelve {dia: [[codigo, razon_social], ...]} para las hojas de día del Excel."""
    tmp = None
    if not path.lower().endswith('.xlsx'):
        tmp = os.path.join(tempfile.gettempdir(), '_gen_rutas_tmp.xlsx')
        shutil.copy(path, tmp)
        path = tmp
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    finally:
        if tmp:
            os.remove(tmp)
    out = {}
    for sn in wb.sheetnames:
        dia = day_of_sheet(sn)
        if not dia:
            continue
        ws = wb[sn]
        maxr = ws.max_row or 0
        maxc = ws.max_column or 0
        hdr = None
        ccol = None
        for r in range(1, min(6, maxr) + 1):
            for c in range(1, maxc + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and 'CODIGO' in norm(v):
                    hdr, ccol = r, c
                    break
            if hdr:
                break
        if not hdr:
            continue
        codes = []
        for r in range(hdr + 1, maxr + 1):
            val = ws.cell(r, ccol).value
            code = None
            if isinstance(val, (int, float)):
                code = str(int(val))
            elif isinstance(val, str) and val.strip().isdigit():
                code = val.strip()
            else:
                s = str(val).strip() if val is not None else ''
                if s.isdigit():
                    code = s
            if code:
                raz = ws.cell(r, ccol + 1).value
                raz = '' if raz is None else str(raz).strip()
                codes.append([code, raz])
        # preserva orden, sin duplicados
        seen = set()
        uniq = []
        for c in codes:
            if c[0] not in seen:
                seen.add(c[0])
                uniq.append(c)
        out[dia] = uniq
    wb.close()
    return out


def main():
    pdv = load_pdv()
    by_code = {p['c']: p for p in pdv}
    rutas = {}
    missing = []
    print('%-38s %-11s %6s  %s' % ('archivo', 'merchan', 'pdv', 'dias'))
    for name, code in FILE_MAP.items():
        path = os.path.join(RUTEOS_DIR, name)
        if not os.path.exists(path):
            print('!! NO EXISTE:', name)
            continue
        dias = extract_codes(path)
        total = 0
        for d in DIAS:
            lst = dias.get(d, [])
            if not lst:
                continue
            total += len(lst)
            for c, _r in lst:
                if c not in by_code:
                    missing.append((code, d, c))
        rutas[code] = dias
        print('%-38s %-11s %6d  %s' % (name, code, total,
                                        ', '.join('%s:%d' % (d, len(dias[d])) for d in DIAS if dias.get(d))))

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(rutas, f, ensure_ascii=False, indent=1)
    print('\nEscribí', OUT)
    print('Codigos de PDV que no estan en pdv.json:')
    for m, d, c in missing:
        print('   %s %s -> %s' % (m, d, c))


if __name__ == '__main__':
    sys.exit(main())
